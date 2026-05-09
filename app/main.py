"""FastAPI service.

Single process serves:
  - REST API at /api/*
  - Static web UI at /
  - Health check at /health

Binds to 0.0.0.0:8765 so a Tailscale-connected phone can reach it
via the laptop's tailnet hostname.
"""
from __future__ import annotations

import io
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from .db import init_db
from . import crud


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    yield


app = FastAPI(title="Tools CRM", version="0.1.0", lifespan=lifespan)

# Permissive CORS so the static page works whether served by us or opened from disk
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------- Schemas ----------

class ContactIn(BaseModel):
    name: str
    company: str | None = None
    role: str | None = None
    phone: str | None = None
    email: str | None = None
    tags: str = ""
    notes: str = ""


class ContactPatch(BaseModel):
    name: str | None = None
    company: str | None = None
    role: str | None = None
    phone: str | None = None
    email: str | None = None
    tags: str | None = None
    notes: str | None = None


class InteractionIn(BaseModel):
    contact_id: int
    channel: str = Field(description="call, email, visit, meeting, other")
    summary: str
    occurred_at: str | None = None
    follow_up_date: str | None = None


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


# ---------- Contacts ----------

@app.get("/api/contacts")
def api_list_contacts(q: str | None = None, tag: str | None = None, limit: int = 200) -> list[dict]:
    return crud.list_contacts(query=q, tag=tag, limit=limit)


@app.get("/api/contacts/{contact_id}")
def api_get_contact(contact_id: int) -> dict:
    contact = crud.get_contact(contact_id)
    if contact is None:
        raise HTTPException(404, "Contact not found")
    contact["interactions"] = crud.list_interactions(contact_id=contact_id)
    return contact


@app.post("/api/contacts", status_code=201)
def api_create_contact(payload: ContactIn) -> dict:
    return crud.create_contact(**payload.model_dump())


@app.patch("/api/contacts/{contact_id}")
def api_update_contact(contact_id: int, payload: ContactPatch) -> dict:
    updated = crud.update_contact(contact_id, **payload.model_dump(exclude_unset=True))
    if updated is None:
        raise HTTPException(404, "Contact not found")
    return updated


@app.delete("/api/contacts/{contact_id}", status_code=204)
def api_delete_contact(contact_id: int) -> None:
    if not crud.delete_contact(contact_id):
        raise HTTPException(404, "Contact not found")


@app.get("/api/tags")
def api_tags() -> list[str]:
    return crud.all_tags()


# ---------- Interactions ----------

@app.get("/api/interactions")
def api_list_interactions(contact_id: int | None = None, limit: int = 100) -> list[dict]:
    return crud.list_interactions(contact_id=contact_id, limit=limit)


@app.post("/api/interactions", status_code=201)
def api_create_interaction(payload: InteractionIn) -> dict:
    if crud.get_contact(payload.contact_id) is None:
        raise HTTPException(404, "Contact not found")
    return crud.create_interaction(**payload.model_dump())


@app.post("/api/interactions/{interaction_id}/complete")
def api_complete_followup(interaction_id: int) -> dict:
    updated = crud.complete_followup(interaction_id)
    if updated is None:
        raise HTTPException(404, "Interaction not found")
    return updated


# ---------- Dashboard ----------

@app.get("/api/today")
def api_today(today: str | None = None, upcoming_days: int = 7) -> dict:
    """Today view: follow-ups due, follow-ups coming up, recent activity."""
    from .db import now_iso
    today_str = today or now_iso()[:10]
    return {
        "today": today_str,
        "followups": crud.todays_followups(today_str),
        "upcoming": crud.upcoming_followups(today_str, days=upcoming_days),
        "recent": crud.recent_activity(days=7),
    }


# ---------- Excel import ----------

@app.post("/api/import/excel")
async def api_import_excel(file: UploadFile = File(...)) -> dict[str, Any]:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Only .xlsx/.xlsm files are supported")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read Excel file: {e}")

    sheet = wb.active
    if sheet is None:
        raise HTTPException(400, "Workbook has no active sheet")

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, "Sheet is empty")

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
    column_map = crud.map_excel_columns(headers)

    if not column_map["name"]:
        raise HTTPException(
            400,
            f"Could not find a name column. Headers found: {headers}. "
            "Expected one of: Name, Navn, Kontaktperson, Contact",
        )

    rows = [dict(zip(headers, r)) for r in rows_iter if any(c is not None for c in r)]
    result = crud.import_contacts_from_rows(rows, column_map, source=file.filename)
    result["column_map"] = column_map
    result["total_rows"] = len(rows)
    return result


# ---------- Excel export ----------

@app.get("/api/export/excel")
def api_export_excel() -> StreamingResponse:
    """Stream a two-sheet workbook (Contacts + Interactions). Contact columns
    are round-trip compatible with the import, so Flip can edit and re-import."""
    contact_rows, interaction_rows = crud.export_rows()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"
    for row in contact_rows:
        ws.append(row)
    ws_i = wb.create_sheet("Interactions")
    for row in interaction_rows:
        ws_i.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="flipbook-{today}.xlsx"'},
    )


# ---------- Static web UI ----------

WEB_DIR = Path(__file__).parent.parent / "web"

if WEB_DIR.exists():
    app.mount("/static", StaticFiles(directory=WEB_DIR), name="static")

    @app.get("/")
    def index() -> FileResponse:
        # no-cache so iterative edits to index.html show up on a normal reload
        # (the page itself is tiny; the CDN-served Alpine + fonts stay cached)
        return FileResponse(
            WEB_DIR / "index.html",
            headers={"Cache-Control": "no-cache, must-revalidate"},
        )
